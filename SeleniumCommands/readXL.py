import openpyxl

path = "C:\\Users\\HP\Documents\\foodchow restaurant manager app.xlsx"

workbook=openpyxl.load_workbook(path)

sheet = workbook.active

rows = sheet.max_row
cols =sheet.max_column

print(rows)
print(cols)

for r in range(11, 20):
    for c in range(1, 10):
        print(sheet.cell(row=r,column=c).value or "         ", end="  " )

    print()
